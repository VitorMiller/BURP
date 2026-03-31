from __future__ import annotations

from typing import Iterable

from openpyxl import load_workbook

from burp.normalization.name import normalize_header


def iter_xlsx_rows(path: str) -> Iterable[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return
    norm_headers = [normalize_header(h) for h in headers]
    for row in rows:
        values = ["" if v is None else str(v).strip() for v in row]
        yield dict(zip(norm_headers, values))
